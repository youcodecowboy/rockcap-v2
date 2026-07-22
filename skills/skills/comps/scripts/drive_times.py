#!/usr/bin/env python3
"""
RockCap comps - proximity columns (Distance + Drive time).

Fills two OPTIONAL columns on a comps Appendix:
  - Distance (mi): exact straight-line from the scheme site to each comp
    (postcodes.io geocode + haversine, no API key needed).
  - Drive (Google Maps, min): real road drive-time from the Google Routes API
    (the same engine behind Google Maps, so the figures match the app).

These columns are NOT standard on every appendix - only populate them when a
scheme wants proximity evidence (Alex's Oakridge Lynch call-note, June 2026).

Why Google Routes, not OSRM: the free no-key OSRM demo server routes on posted
speed limits with no traffic and runs ~20-40% short of Google on rural roads.
Google Routes returns Google's own numbers. Straight-line distance is exact
either way.

SETUP (one-time): create a Google Maps Platform API key with the **Routes API**
enabled on the firm's existing Google Cloud project, restrict it to the Routes
API, and add to ~/.claude/.env as:
    GOOGLE_MAPS_API_KEY=...
At RockCap comp volumes (10-30 comps/scheme) this sits inside Google's monthly
free tier. Verify current pricing when minting the key.

USAGE:
    # programmatic
    from drive_times import populate_appendix
    populate_appendix("path/to/Appendix.xlsx", origin="Oakridge Lynch GL6 7NR",
                      dist_col=11, drive_col=12)

    # CLI
    python3 drive_times.py "path/to/Appendix.xlsx" "Oakridge Lynch GL6 7NR"

Reads comps from rows that have a location/postcode in column B and a price in
column D (skips tier headers, subject units, summary/conclusion rows). Subject
units (col B == the scheme name only) are set to 0 mi / "On site".

Drive-time accuracy note: routingPreference defaults to TRAFFIC_UNAWARE, a
stable typical figure (no live-traffic swing) - best for a document that must
read the same next week. Switch TRAFFIC_PREF to "TRAFFIC_AWARE" with a
departure_time if a live-traffic figure is ever wanted.
"""
import json, math, os, re, subprocess, sys, time, urllib.parse, urllib.request

POSTCODES_IO = "https://api.postcodes.io/postcodes/"
ROUTES_URL = "https://routes.googleapis.com/distanceMatrix/v2:computeRouteMatrix"
PC_FULL = re.compile(r"[A-Z]{1,2}[0-9][A-Z0-9]?\s*[0-9][A-Z]{2}")
TRAFFIC_PREF = "TRAFFIC_UNAWARE"  # stable typical time; "TRAFFIC_AWARE" for live


def _load_env_key():
    """Read GOOGLE_MAPS_API_KEY from ~/.claude/.env (single source of truth)."""
    key = os.environ.get("GOOGLE_MAPS_API_KEY")
    if key:
        return key
    env = os.path.expanduser("~/.claude/.env")
    if os.path.exists(env):
        with open(env) as f:
            for line in f:
                line = line.strip()
                if line.startswith("GOOGLE_MAPS_API_KEY="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
    return None


def _get_json(url):
    with urllib.request.urlopen(url, timeout=25) as r:
        return json.load(r)


def geocode(text):
    """Return (lat, lon) for a UK postcode found in `text`, else None."""
    m = PC_FULL.search(str(text).upper().replace("  ", " "))
    if not m:
        return None
    pc = m.group(0).replace(" ", "")
    try:
        r = _get_json(POSTCODES_IO + urllib.parse.quote(pc))["result"]
        return (r["latitude"], r["longitude"])
    except Exception:
        return None


def haversine_mi(a, b):
    R = 3958.8
    dlat = math.radians(b[0] - a[0])
    dlon = math.radians(b[1] - a[1])
    x = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(a[0])) * math.cos(math.radians(b[0]))
         * math.sin(dlon / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(x))


def _waypoint(x):
    """Accept an (lat, lon) tuple -> latLng waypoint, or a str -> address."""
    if isinstance(x, (tuple, list)):
        return {"waypoint": {"location": {"latLng": {"latitude": x[0],
                                                     "longitude": x[1]}}}}
    return {"waypoint": {"address": x}}


def google_drive_minutes(origin, dests, api_key):
    """
    origin / each dest: either an (lat, lon) tuple (preferred - exact, no
    name-ambiguity) or an address string. Returns list of drive-times in
    minutes (int), aligned to dests; None per element on failure. Uses
    computeRouteMatrix (one origin -> many dests).

    Always prefer coordinates: routing on a property-name string lets Google
    geocode to the wrong same-named property (e.g. "St Michaels Cottage"
    matched a different one and returned 8 min for a 0.2mi neighbour). Passing
    the postcode-geocoded latLng pins the right location.
    """
    body = {
        "origins": [_waypoint(origin)],
        "destinations": [_waypoint(d) for d in dests],
        "travelMode": "DRIVE",
        "routingPreference": TRAFFIC_PREF,
    }
    headers = [
        "-H", "Content-Type: application/json",
        "-H", "X-Goog-Api-Key: " + api_key,
        "-H", "X-Goog-FieldMask: originIndex,destinationIndex,duration,condition",
    ]
    # curl (Python urllib hits a TLS handshake failure against some Google
    # routing endpoints; curl is reliable).
    cmd = ["curl", "-s", "--max-time", "30", "-X", "POST", ROUTES_URL,
           *headers, "-d", json.dumps(body)]
    out = subprocess.run(cmd, capture_output=True, text=True).stdout
    res = [None] * len(dests)
    try:
        data = json.loads(out)
    except Exception:
        print("Routes API parse error:", out[:300], file=sys.stderr)
        return res
    if isinstance(data, dict) and data.get("error"):
        print("Routes API error:", data["error"].get("message"), file=sys.stderr)
        return res
    for el in data:
        di = el.get("destinationIndex")
        dur = el.get("duration")  # e.g. "612s"
        if di is None or not dur:
            continue
        if el.get("condition") not in (None, "ROUTE_EXISTS"):
            continue
        res[di] = round(int(str(dur).rstrip("s")) / 60)
    return res


def populate_appendix(path, origin, dist_col=11, drive_col=12,
                      origin_coord=None, scheme_name=None):
    """
    Fill Distance (straight-line mi) + Drive (Google min) columns in a comps
    Appendix .xlsx. Rebuilds nothing structural - only writes the two columns,
    so it is safe to run on a finished, hyperlinked sheet.

    origin: address/postcode string for routing (e.g. "Oakridge Lynch GL6 7NR").
    origin_coord: optional (lat, lon) for the straight-line origin; derived from
        `origin` postcode if omitted.
    scheme_name: subject rows (col B == this) are set 0 mi / "On site".
    """
    import openpyxl
    api_key = _load_env_key()
    if not api_key:
        raise SystemExit(
            "No GOOGLE_MAPS_API_KEY in ~/.claude/.env - create a Maps Platform "
            "key with the Routes API enabled and add it. See module docstring.")
    if origin_coord is None:
        origin_coord = geocode(origin)
        if origin_coord is None:
            raise SystemExit("Could not geocode origin postcode: " + origin)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    if scheme_name is None:
        scheme_name = (origin.split(" GL")[0]
                       if " GL" in origin else origin).strip()

    SKIP = ("Tier", "Address", "Subject", "Set", "Type", "MARKET", "Headline",
            "SIZE", "CONCLUSION", "AFFORDABLE", "Local", "Prime ")
    comp_rows, dests = [], []
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "")
        b = ws.cell(r, 2).value
        d = ws.cell(r, 4).value
        if not (a and isinstance(d, (int, float))):
            continue
        if a.startswith(SKIP):
            continue
        if str(b).strip() == scheme_name:                      # subject unit
            ws.cell(r, dist_col, 0).number_format = "0.0"
            ws.cell(r, drive_col, "On site")
            continue
        coord = geocode(b)
        if coord:
            ws.cell(r, dist_col, round(haversine_mi(origin_coord, coord), 1))
            ws.cell(r, dist_col).number_format = "0.0"
            comp_rows.append(r)
            dests.append(coord)                 # route on coords, not name
        else:                                    # no full postcode -> can't pin
            ws.cell(r, drive_col, "n/a")

    if dests:
        mins = google_drive_minutes(origin_coord, dests, api_key)
        for r, m in zip(comp_rows, mins):
            ws.cell(r, drive_col, m if m is not None else "n/a")

    wb.save(path)
    filled = sum(1 for r in comp_rows
                 if isinstance(ws.cell(r, drive_col).value, (int, float)))
    print(f"populated {len(comp_rows)} comps; drive-times filled "
          f"{filled}/{len(comp_rows)} (Google Routes, {TRAFFIC_PREF})")
    return filled, len(comp_rows)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print('usage: python3 drive_times.py <appendix.xlsx> "<origin postcode>" '
              '[dist_col] [drive_col]')
        sys.exit(1)
    p, origin = sys.argv[1], sys.argv[2]
    dc = int(sys.argv[3]) if len(sys.argv) > 3 else 11
    lc = int(sys.argv[4]) if len(sys.argv) > 4 else 12
    populate_appendix(p, origin, dist_col=dc, drive_col=lc)
