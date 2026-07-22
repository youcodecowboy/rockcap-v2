#!/usr/bin/env python3.12
"""
externalise_model.py  —  RockCap internal appraisal/portfolio .xlsm  ->  client/lender EXTERNAL .xlsm

WHY THIS EXISTS (read before touching):
  RockCap appraisal models carry the RockCap logo (an embedded picture) on every sheet.
  openpyxl — the obvious library — SILENTLY DROPS embedded images on save, so a rebuild loses
  every logo and re-serialises formatting. The ONLY way to keep logos + formatting EXACTLY is
  to operate on the file at the zip/XML level: copy the source's worksheet/drawing/media/chart/
  VBA parts verbatim and only (a) delete unwanted tabs, (b) "break links" by freezing formulas
  to cached values, (c) rename + reorder kept tabs, (d) clean errors, drawings and metadata.

  This script never uses openpyxl to WRITE. It uses openpyxl read-only ONLY to read scheme
  names and to run the final open-test. All transformation is raw-zip surgery.

  These models all come from one Excel template, so the worksheet XML shape is consistent
  (double-quoted attributes, unprefixed namespaces, UTF-8). The regexes assume that shape;
  if a future model is authored differently, prefer a structural XML parse.

USAGE:
  python3.12 externalise_model.py SOURCE.xlsm OUT.xlsm
      [--dashboard portfolio|lender]   # portfolio = "Portfolio Dashboard - BFS" (default, all
                                       #   active sites); lender = "Lender Dashboard - BFS" +
                                       #   site 1 only (single-scheme model)
      [--no-consol]                    # drop "Consol Cashflow"
      [--keep "Exact Tab,Exact Tab"]   # OVERRIDE auto-detect: keep exactly these (no rename)
      [--clear-rows "Sheet:209,246;Other:33"]  # blank whole rows (model-specific obsolete blocks)
      [--no-fix-title]                 # do NOT rewrite a "...- BTR" dashboard title to "...- BFS"
      [--keep-orphans]                 # do NOT strip drawings/charts of deleted sheets

  Requires only the Python stdlib + openpyxl (read-only name lookup). No Pillow.
"""
import argparse, re, sys, zipfile
from xml.sax.saxutils import escape, unescape

# Canonical internal-only tab names. Deletion is driven by "not in the keep set" (see surgery());
# this list is used by verify() to assert none of these names survive in tabs or metadata.
INTERNAL_TABS = {
    "Control Sheet", "Central Inputs", "Categories",
    "Portfolio Dashboard - BTR", "Lender Dashboard - BTR", "Lender Dashboard - BFS",
    "Sensitivity Testing - BFS", "Sensitivity Testing - BTR",
    "Lender BTR Dashboard Workings", "Lender BFS Dashboard Workings",
}
# A site whose Project Name (C9) is blank/one of these is an unused template tab — never keep it.
# "Project <n>" (digits only) is the model's default site name for unfilled tabs — matched
# precisely so a genuinely-named scheme ("Project Blackfriars") is never skipped.
PLACEHOLDER = re.compile(r"^\s*(spare|example|enter|tbc|placeholder|n/?a|project\s*\d+\s*$)", re.I)
ILLEGAL_TAB = re.compile(r"[\\/?*\[\]:]")  # Excel-forbidden tab-name characters


def sanitize_tab(name):
    """Make a string a valid Excel tab name: strip forbidden chars, collapse spaces, cap 31."""
    n = ILLEGAL_TAB.sub(" ", name).strip()
    n = re.sub(r"\s+", " ", n)
    return (n[:31].rstrip() or "Sheet")


def read_model(src):
    """Read-only: return (ordered sheet names, {AppraisalSiteN: scheme name from C9})."""
    import openpyxl
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    sheets = list(wb.sheetnames)
    schemes = {}
    for s in sheets:
        if re.fullmatch(r"AppraisalSite(\d+)", s):
            v = wb[s]["C9"].value  # RockCap convention: C9 = Project Name
            schemes[s] = (str(v).strip() if v is not None else "")
    wb.close()
    return sheets, schemes


def build_spec(sheets, schemes, dashboard, keep_consol, keep_override):
    """Ordered list of (orig_name, new_name) to keep. Skips placeholder/blank sites, sanitises
    and de-duplicates final names — Excel forbids duplicate or >31-char tab names, so a portfolio
    where two sites share a Project Name would otherwise produce a workbook that will not open."""
    if keep_override:
        return [(n, n) for n in keep_override]
    order = []
    dash = "Portfolio Dashboard - BFS" if dashboard == "portfolio" else "Lender Dashboard - BFS"
    if dash in sheets:
        order.append((dash, dash))
    elif dashboard == "portfolio" and "Lender Dashboard - BFS" in sheets:
        order.append(("Lender Dashboard - BFS", "Lender Dashboard - BFS"))
    if keep_consol and "Consol Cashflow" in sheets:
        order.append(("Consol Cashflow", "Consol Cashflow"))
    idxs = sorted(int(re.fullmatch(r"AppraisalSite(\d+)", s).group(1))
                  for s in sheets if re.fullmatch(r"AppraisalSite(\d+)", s))
    if dashboard == "lender":
        idxs = idxs[:1]  # single-scheme model -> site 1 only
    seen, skipped = {}, []
    for i in idxs:
        a, c = f"AppraisalSite{i}", f"CashflowSite{i}"
        scheme = (schemes.get(a) or "").strip()
        if not scheme or PLACEHOLDER.match(scheme):
            skipped.append(f"Site{i} ({scheme or 'blank'})")
            continue
        for tab, role in ((a, "Appraisal"), (c, "Cashflow")):
            if tab not in sheets:
                continue
            base = sanitize_tab(f"{scheme} {role}")
            n = seen.get(base, 0) + 1
            seen[base] = n
            name = base if n == 1 else sanitize_tab(f"{base} ({n})")
            order.append((tab, name))
    if skipped:
        print("Skipped placeholder/blank sites:", ", ".join(skipped))
    return order


def freeze_xml(xml, clear_rows=None):
    """Break links: strip <f> (self-closing FIRST), blank error cells, drop data validations,
    clear obsolete rows. Keeps cached <v> values so visible results survive."""
    # ORDER MATTERS. Shared-formula cells are <f t="shared" si="0"/> (self-closing). If the
    # paired regex runs first, its [^>]* swallows the trailing slash and .*?</f> races to the
    # NEXT </f>, deleting every cell in between. Remove self-closing <f/> first, then paired.
    xml = re.sub(r"<f\b[^>]*?/>", "", xml)
    xml = re.sub(r"<f\b[^>]*?>.*?</f>", "", xml, flags=re.DOTALL)
    # Blank ALL error cells (t="e": #REF! #DIV/0! #N/A #VALUE! #NAME? #NUM! #NULL!). Preserve
    # every other attribute (incl. the style s="...") whether it sits before or after t="e".
    xml = re.sub(r'<c\b([^>]*?)\st="e"([^>]*?)>\s*<v>[^<]*</v>\s*</c>', r"<c\1\2/>", xml)
    # Drop data validations — a view-only file needs no input dropdowns, and a DV that pointed at
    # a now-deleted sheet/range would be a dangling reference.
    xml = re.sub(r"<dataValidations\b.*?</dataValidations>", "", xml, flags=re.DOTALL)
    xml = re.sub(r"<x14:dataValidations\b.*?</x14:dataValidations>", "", xml, flags=re.DOTALL)
    # Clear specific obsolete rows entirely (label + data), keeping the row's own attributes.
    for r in (clear_rows or []):
        xml = re.sub(rf'(<row\b[^>]*?\br="{r}"[^>]*?)>.*?</row>', r"\1/>", xml, flags=re.DOTALL)
    return xml


def referenced_parts(parts, kept_sheet_parts):
    """Set of drawing/chart/vml/media parts reachable from the KEPT sheets (transitively)."""
    keep = set()

    def follow(part):
        rels = part.rsplit("/", 1)
        relpath = rels[0] + "/_rels/" + rels[1] + ".rels"
        if relpath not in parts:
            return
        base = part.rsplit("/", 1)[0]
        for tgt in re.findall(r'Target="([^"]+)"', parts[relpath].decode("utf8", "ignore")):
            if tgt.startswith("http") or "vbaProject" in tgt:
                continue
            norm = tgt
            # resolve ../ relative to the part's folder
            cur = base.split("/")
            for seg in tgt.split("/"):
                if seg == "..":
                    cur = cur[:-1]
                else:
                    cur = cur + [seg]
            norm = "/".join(cur).replace("xl/xl/", "xl/")
            if norm not in keep and norm in parts:
                keep.add(norm)
                follow(norm)
    for sp in kept_sheet_parts:
        follow(sp)
    return keep


def surgery(src, out, spec, clear_map, fix_title, strip_orphans):
    with zipfile.ZipFile(src) as z:
        parts = {n: z.read(n) for n in z.namelist()}

    wb = parts["xl/workbook.xml"].decode("utf8")
    rels = parts["xl/_rels/workbook.xml.rels"].decode("utf8")
    ct = parts["[Content_Types].xml"].decode("utf8")

    # name -> (sheetId, r:id); names are XML-unescaped to match openpyxl's view
    name2 = {}
    for s in re.findall(r"<sheet\b[^>]*/>", wb):
        nm = unescape(re.search(r'name="([^"]*)"', s).group(1))
        sid = re.search(r'sheetId="(\d+)"', s).group(1)
        rid = re.search(r'r:id="(rId\d+)"', s).group(1)
        name2[nm] = (sid, rid)
    rid2part = {rid: ("xl/" + tgt.lstrip("/").replace("xl/", "", 1) if tgt.startswith("/") else "xl/" + tgt)
                for rid, tgt in re.findall(r'Id="(rId\d+)"[^>]*?Target="([^"]*worksheets/sheet\d+\.xml)"', rels)}

    keep = {o for o, _ in spec}
    deleted = [nm for nm in name2 if nm not in keep]

    del_parts = set()
    for nm in deleted:
        part = rid2part[name2[nm][1]]
        del_parts.add(part)
        del_parts.add(part.replace("worksheets/", "worksheets/_rels/") + ".rels")
    del_parts.add("xl/calcChain.xml")

    # rebuild <sheets> in target order with renamed display names (sheetId/r:id unchanged)
    new_sheets = "<sheets>" + "".join(
        f'<sheet name="{escape(new)}" sheetId="{name2[orig][0]}" r:id="{name2[orig][1]}"/>'
        for orig, new in spec) + "</sheets>"
    wb = re.sub(r"<sheets>.*?</sheets>", lambda _: new_sheets, wb, flags=re.DOTALL)
    wb = re.sub(r"<definedNames>.*?</definedNames>", "", wb, flags=re.DOTALL)

    del_rids = {name2[nm][1] for nm in deleted}
    def keep_rel(m):
        r = m.group(0)
        rid = re.search(r'Id="(rId\d+)"', r)
        return "" if (rid and rid.group(1) in del_rids) or "calcChain.xml" in r else r
    rels = re.sub(r"<Relationship\b[^>]*/>", keep_rel, rels)

    kept_sheet_parts = {rid2part[name2[o][1]] for o, _ in spec}

    # strip orphan drawings/charts/vml of deleted sheets (they hold cached values + shapes; the
    # worksheet data is already gone, but orphan chart caches still leak on inspection)
    if strip_orphans:
        keep_aux = referenced_parts(parts, kept_sheet_parts)
        for n in list(parts):
            if re.match(r"xl/(drawings|charts)/.*\.(xml|vml)$", n) and n not in keep_aux:
                del_parts.add(n)
                rp = n.rsplit("/", 1)
                del_parts.add(rp[0] + "/_rels/" + rp[1] + ".rels")

    # content types: drop overrides for every removed .xml part
    del_pn = {"/" + p for p in del_parts if p.endswith(".xml")}
    ct = re.sub(r"<Override\b[^>]*/>",
                lambda m: "" if re.search(r'PartName="([^"]*)"', m.group(0)).group(1) in del_pn else m.group(0),
                ct)

    parts["xl/workbook.xml"] = wb.encode("utf8")
    parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf8")
    parts["[Content_Types].xml"] = ct.encode("utf8")

    # freeze + blank-errors + drop-DV + clear-rows on every KEPT sheet
    for orig, _ in spec:
        part = rid2part[name2[orig][1]]
        parts[part] = freeze_xml(parts[part].decode("utf8"), clear_map.get(orig, [])).encode("utf8")

    # fix a "...- BTR" dashboard title that mislabels the kept BFS dashboard. Do it on BOTH the
    # sheet XML (inline strings) AND sharedStrings (the usual home) so no stale "BTR" string is
    # left to leak. Safe because the real BTR dashboard has been deleted.
    if fix_title and any("Portfolio Dashboard - BFS" == n for _, n in spec):
        for tgt in ("xl/sharedStrings.xml", rid2part[name2["Portfolio Dashboard - BFS"][1]]):
            if tgt in parts:
                parts[tgt] = parts[tgt].replace(b"Portfolio Dashboard - BTR",
                                                b"Portfolio Dashboard - BFS")

    # docProps/custom.xml — Microsoft sensitivity labels etc.; must not go external
    if "docProps/custom.xml" in parts:
        del parts["docProps/custom.xml"]
        parts["[Content_Types].xml"] = re.sub(
            r'<Override PartName="/docProps/custom\.xml"[^>]*/>', "",
            parts["[Content_Types].xml"].decode("utf8")).encode("utf8")
        for rp in ("_rels/.rels", "docProps/_rels/app.xml.rels"):
            if rp in parts:
                parts[rp] = re.sub(r"<Relationship[^>]*custom\.xml[^>]*/>", "",
                                   parts[rp].decode("utf8")).encode("utf8")

    # docProps/app.xml — rebuild so deleted tab names + the old named-range count don't leak
    if "docProps/app.xml" in parts:
        titles = [new for _, new in spec]
        hp = ('<HeadingPairs><vt:vector size="2" baseType="variant"><vt:variant>'
              '<vt:lpstr>Worksheets</vt:lpstr></vt:variant><vt:variant>'
              f'<vt:i4>{len(titles)}</vt:i4></vt:variant></vt:vector></HeadingPairs>')
        tp = (f'<TitlesOfParts><vt:vector size="{len(titles)}" baseType="lpstr">'
              + "".join(f"<vt:lpstr>{escape(t)}</vt:lpstr>" for t in titles)
              + "</vt:vector></TitlesOfParts>")
        ax = parts["docProps/app.xml"].decode("utf8")
        ax = re.sub(r"<HeadingPairs>.*?</HeadingPairs>", lambda _: hp, ax, flags=re.DOTALL)
        ax = re.sub(r"<TitlesOfParts>.*?</TitlesOfParts>", lambda _: tp, ax, flags=re.DOTALL)
        parts["docProps/app.xml"] = ax.encode("utf8")

    # scrub author metadata
    if "docProps/core.xml" in parts:
        c = parts["docProps/core.xml"].decode("utf8")
        c = re.sub(r"<dc:creator>.*?</dc:creator>", "<dc:creator>RockCap</dc:creator>", c)
        c = re.sub(r"<cp:lastModifiedBy>.*?</cp:lastModifiedBy>",
                   "<cp:lastModifiedBy>RockCap</cp:lastModifiedBy>", c)
        parts["docProps/core.xml"] = c.encode("utf8")

    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zo:
        for n, b in parts.items():
            if n not in del_parts:
                zo.writestr(n, b)
    return deleted


def verify(out, keep_final=()):
    """Pass/fail report. Gates on the failure modes that have actually bitten us."""
    keep_final = set(keep_final)
    with zipfile.ZipFile(out) as z:
        names = z.namelist()
        ws = [n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n)]
        # cell formula tags only (<f>, <f ...>, <f/>); NOT <formula> in data validations
        formulas = sum(len(re.findall(rb"<f[\s/>]", z.read(n))) for n in ws)
        errors = sum(z.read(n).count(b't="e"') for n in ws)
        logos = sum(z.read(n).count(b"<xdr:pic>") for n in names if re.match(r"xl/drawings/drawing\d+\.xml$", n))
        media = len([n for n in names if "xl/media/" in n])
        vba = "xl/vbaProject.bin" in names
        wbxml = z.read("xl/workbook.xml")
        macro_ct = b"macroEnabled.main+xml" in z.read("[Content_Types].xml")
        defined = len(re.findall(rb"<definedName\b", wbxml))
        ext_links = any("externalLink" in n or "pivotCache" in n or "connections.xml" in n for n in names)
        app = z.read("docProps/app.xml") if "docProps/app.xml" in names else b""
        core = z.read("docProps/core.xml") if "docProps/core.xml" in names else b""
        blob = b"".join(z.read(n) for n in names)
        sstr = z.read("xl/sharedStrings.xml") if "xl/sharedStrings.xml" in names else b""
        leaked_meta = sorted(t for t in INTERNAL_TABS
                             if t not in keep_final and (t.encode() in app or t.encode() in core))
        leaked_str = sorted(t for t in INTERNAL_TABS if t not in keep_final and t.encode() in sstr)
        sensitivity = (b"Unclassified" in blob or b"MSIP_Label" in blob
                       or b"sensitivitylabel" in blob.lower())
    import openpyxl
    tabs = openpyxl.load_workbook(out).sheetnames  # raises on a malformed package
    long_tabs = [t for t in tabs if len(t) > 31]
    dup = len(tabs) != len(set(tabs))
    gate_fail = (formulas or errors or not vba or not macro_ct or leaked_meta
                 or sensitivity or defined or ext_links or long_tabs or dup)
    print("\n=== VERIFY ===")
    print(f"  opens in openpyxl : OK")
    print(f"  tabs ({len(tabs)})       : {tabs[:4]} ...")
    print(f"  duplicate / >31ch : {dup} / {long_tabs or 'none'}   (want False / none)")
    print(f"  formulas left     : {formulas}   (want 0)")
    print(f"  error cells (t=e) : {errors}   (want 0)")
    print(f"  defined names     : {defined}   (want 0)")
    print(f"  external links    : {ext_links}   (want False)")
    print(f"  logos <xdr:pic>   : {logos}   (kept-sheet logos present)")
    print(f"  media files       : {media}   (1 = standard RockCap logo; >1 only if a 2nd real image)")
    print(f"  vbaProject / .xlsm: {vba} / macroCT={bool(macro_ct)}   (both True -> opens clean)")
    print(f"  metadata leak     : {('LEAK ' + str(leaked_meta)) if leaked_meta else 'clean'}")
    if leaked_str:
        print(f"  WARNING: internal names still in sharedStrings (orphan, non-fatal): {leaked_str}")
    print(f"  sensitivity label : {'PRESENT' if sensitivity else 'none'}")
    print(f"  RESULT            : {'CHECK ABOVE' if gate_fail else 'PASS'}")
    return not gate_fail


def main():
    ap = argparse.ArgumentParser(description="RockCap internal model -> EXTERNAL view-only workbook")
    ap.add_argument("source")
    ap.add_argument("out")
    ap.add_argument("--dashboard", choices=["portfolio", "lender"], default="portfolio")
    ap.add_argument("--no-consol", action="store_true")
    ap.add_argument("--keep", default="", help="override: exact tab names, comma-separated")
    ap.add_argument("--clear-rows", default="", help='e.g. "Consol Cashflow:209,246;Goring Appraisal:33"')
    ap.add_argument("--no-fix-title", action="store_true")
    ap.add_argument("--keep-orphans", action="store_true", help="keep drawings/charts of deleted sheets")
    a = ap.parse_args()

    import os
    if not os.path.isfile(a.source):
        ap.error(f"source not found: {a.source}")
    if not a.out.lower().endswith(".xlsm"):
        ap.error("output must be .xlsm (keeps the VBA project so it opens cleanly)")

    sheets, schemes = read_model(a.source)
    keep_override = [s.strip() for s in a.keep.split(",") if s.strip()] if a.keep else None
    if keep_override:
        missing = [k for k in keep_override if k not in sheets]
        if missing:
            ap.error(f"--keep names not in workbook: {missing}")
    spec = build_spec(sheets, schemes, a.dashboard, not a.no_consol, keep_override)
    if len(spec) < 2:
        ap.error("nothing to keep — check --dashboard / --keep and that the model has site tabs")

    clear_map = {}
    for chunk in [c for c in a.clear_rows.split(";") if c.strip()]:
        if ":" not in chunk:
            ap.error(f'--clear-rows chunk needs Sheet:rows form, got "{chunk}"')
        sheet, rows = chunk.split(":", 1)
        clear_map[sheet.strip()] = [int(r) for r in rows.split(",") if r.strip()]

    print("Source sheets:", len(sheets), "| schemes detected:", len(schemes))
    print("Keeping", len(spec), "tabs:")
    for o, n in spec:
        print(f"   {o:28} -> {n}")
    deleted = surgery(a.source, a.out, spec, clear_map, not a.no_fix_title, not a.keep_orphans)
    print("Deleted", len(deleted), "internal tabs:", deleted)
    ok = verify(a.out, {new for _, new in spec})
    print(f"\nWrote: {a.out}")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
