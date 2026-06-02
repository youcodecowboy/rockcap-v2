"""
Export prospect candidates for a lender to XLSX, enriched with Companies House
basic profile info.

This is the "sourcing" enrichment in miniature: take the company numbers a lender
has charged, fetch each company's basic CH profile (one call -> name, status,
type, SIC, registered office), and write a client-ready spreadsheet that carries
the provenance (which lender, which scheme) alongside.

One CH profile call returns an expanded record, not just the name. The CH API
allows ~600 requests / 5 min per key, so a few hundred candidates is well within
budget; we throttle to be polite since the key is shared with production.

Usage:
    DATABASE_URL=... COMPANIES_HOUSE_API_KEY=... \
    python -m etl.export_candidates \
        --lender "PARAGON DEVELOPMENT FINANCE LIMITED" \
        --since 2024-01-01 --out ~/ch-bulk/paragon_candidates.xlsx
"""

from __future__ import annotations

import argparse
import os
import time
from base64 import b64encode
from urllib.request import Request, urlopen
import json

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CH_BASE = "https://api.company-information.service.gov.uk"


def fetch_candidates(dsn: str, lender: str, since: str | None, status: str):
    """Aggregate the charge facts per company for a lender."""
    clause = "WHERE lender_canonical = %s"
    params: list = [lender.upper()]
    if since:
        clause += " AND date_registered >= %s"
        params.append(since)
    if status == "outstanding":
        clause += " AND NOT fully_satisfied"
    elif status == "satisfied":
        clause += " AND fully_satisfied"
    sql = f"""
        SELECT company_number,
               COUNT(*)                                       AS charge_count,
               MAX(date_registered)                           AS latest_charge,
               BOOL_OR(NOT fully_satisfied)                   AS has_outstanding,
               (array_agg(property ORDER BY date_registered DESC NULLS LAST))[1] AS recent_property
        FROM charges
        {clause}
        GROUP BY company_number
        ORDER BY latest_charge DESC NULLS LAST
    """
    with psycopg.connect(dsn) as c:
        return c.execute(sql, params).fetchall()


def ch_profile(number: str, auth_header: str) -> dict:
    req = Request(f"{CH_BASE}/company/{number}", headers={"Authorization": auth_header})
    try:
        with urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except Exception:
        return {}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--lender", required=True)
    ap.add_argument("--since")
    ap.add_argument("--status", default="all", choices=["all", "outstanding", "satisfied"])
    ap.add_argument("--out", required=True)
    ap.add_argument("--delay", type=float, default=0.4, help="seconds between CH calls (rate limit)")
    args = ap.parse_args()

    dsn = os.environ["DATABASE_URL"]
    key = os.environ["COMPANIES_HOUSE_API_KEY"]
    auth = "Basic " + b64encode(f"{key}:".encode()).decode()

    rows = fetch_candidates(dsn, args.lender, args.since, args.status)
    print(f"{len(rows)} candidate companies; enriching via Companies House...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sourcing Candidates"
    headers = [
        "Company Name", "Company Number", "Status", "Type", "Incorporated",
        "SIC Code(s)", "Town", "Postcode", "Sourced From (Lender)",
        "Charges", "Latest Charge", "Outstanding", "Most Recent Scheme / Property",
    ]
    ws.append(headers)
    header_fill = PatternFill(patternType="solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for i, (num, count, latest, outstanding, prop) in enumerate(rows, 1):
        p = ch_profile(num, auth)
        addr = p.get("registered_office_address", {}) or {}
        ws.append([
            p.get("company_name", "(not found)"),
            num,
            p.get("company_status", ""),
            p.get("type", ""),
            p.get("date_of_creation", ""),
            ", ".join(p.get("sic_codes", []) or []),
            addr.get("locality", ""),
            addr.get("postal_code", ""),
            args.lender,
            count,
            latest.isoformat() if latest else "",
            "Yes" if outstanding else "No",
            (prop or "")[:120],
        ])
        if i % 50 == 0:
            print(f"  ...{i}/{len(rows)}")
        time.sleep(args.delay)

    # Reasonable column widths + freeze header.
    widths = [34, 15, 10, 8, 13, 14, 18, 11, 34, 9, 13, 12, 50]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    wb.save(args.out)
    print(f"wrote {args.out}")


if __name__ == "__main__":
    main()
